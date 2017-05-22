#coding: utf-8
# 依赖openpyxl库：http://openpyxl.readthedocs.org/en/latest/

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook
import csv
import os
import sys

def xlsx2csv(filename):
    try:
        xlsx_file_reader = load_workbook(filename=filename)
        for sheet in xlsx_file_reader.get_sheet_names():
            # 每个sheet输出到一个csv文件中，文件名用xlsx文件名和sheet名用'_'连接
            csv_filename = '{xlsx}.csv'.format(
                xlsx=os.path.splitext(filename.replace(' ', '_'))[0])

            csv_file = file(csv_filename, 'wb')
            csv_file_writer = csv.writer(csv_file)

            sheet_ranges = xlsx_file_reader[sheet]
            for row in sheet_ranges.rows:
                row_container = []
                for cell in row:
                    if type(cell.value) == unicode:
                        row_container.append(cell.value.encode('utf-8'))
                    else:
                        row_container.append(str(cell.value))
                csv_file_writer.writerow(row_container)
            csv_file.close()

    except Exception as e:
        print(e)


def get_all_xlsx(rootDir):
    filenames = {}
    for name in os.listdir(rootDir):
        path = os.path.join(rootDir, name)
        if name.endswith('.xlsx'):
            attr_name = path.split('\\')[-1].split('.')[0]
            filenames[attr_name] = path
        elif os.path.isdir(path):
            get_all_xlsx(path)
    return filenames


if __name__ == '__main__':
    filenames = get_all_xlsx('.')
    for f in filenames.itervalues():
        xlsx2csv(f)
    print 'ok'
